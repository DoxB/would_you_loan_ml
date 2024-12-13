import pandas as pd
import numpy as np
import joblib
import logging
from sklearn.feature_extraction.text import TfidfVectorizer
import openpyxl

# 로깅 설정
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def load_model_and_vectorizer(model_path: str, vectorizer_path: str):
    """모델과 벡터라이저 로드"""
    try:
        model = joblib.load(model_path)
        vectorizer = joblib.load(vectorizer_path)
        logger.info("모델과 벡터라이저를 성공적으로 로드했습니다.")
        return model, vectorizer
    except Exception as e:
        logger.error(f"모델 또는 벡터라이저 로딩 중 오류 발생: {str(e)}")
        raise

def prepare_text(row):
    """텍스트 전처리"""
    title = str(row['제목']).strip()
    keywords = str(row['키워드']).strip() if pd.notna(row['키워드']) else ''
    features = str(row['특성추출(가중치순 상위 50개)']).strip() if pd.notna(row['특성추출(가중치순 상위 50개)']) else ''
    return f"{title} {keywords} {features}"

def analyze_sentiments(df: pd.DataFrame, model, vectorizer):
    """감성 분석 수행"""
    try:
        # 텍스트 전처리
        processed_texts = df.apply(prepare_text, axis=1)
        logger.info(f"전처리된 텍스트 수: {len(processed_texts)}")

        # 벡터라이저 변환
        transformed_texts = vectorizer.transform(processed_texts)
        logger.info(f"벡터라이저 변환 shape: {transformed_texts.shape}")

        # 예측 수행
        predictions = model.predict(transformed_texts)
        predictions = predictions.astype(int)
        
        logger.info(f"예측 완료. Unique 예측값: {np.unique(predictions)}")
        logger.info(f"예측값 분포:\n{pd.Series(predictions).value_counts()}")
        
        return predictions

    except Exception as e:
        logger.error(f"감성 분석 중 오류 발생: {str(e)}")
        raise

# def filter_top_n_by_date(df: pd.DataFrame, sentiment_labels: np.ndarray, n=5):
#     """일자별로 최대 n개의 긍정, 부정, 중립 데이터 필터링"""
#     try:
#         # 감성 라벨을 데이터프레임에 추가
#         df['sentiment_label'] = sentiment_labels
#         df['sentiment_desc'] = df['sentiment_label'].map({1: '긍정', 0: '중립', -1: '부정'})

#         # 일자별로 긍정, 부정, 중립 데이터를 필터링
#         filtered_data = []
#         grouped = df.groupby('일자')
#         for date, group in grouped:
#             positive_samples = group[group['sentiment_desc'] == '긍정'].head(n)
#             negative_samples = group[group['sentiment_desc'] == '부정'].head(n)
#             neutral_samples = group[group['sentiment_desc'] == '중립'].head(max(1, int(n * 0.2)))
#             filtered_data.append(positive_samples)
#             filtered_data.append(negative_samples)
#             filtered_data.append(neutral_samples)
        
    #     # 결과 합치기
    #     result_df = pd.concat(filtered_data)
    #     return result_df.reset_index(drop=True)
    # except Exception as e:
    #     logger.error(f"일자별 데이터 필터링 중 오류 발생: {str(e)}")
    #     raise

def main():
    try:
        # 파일 경로 설정
        input_file = "NewsResult_20201205-20241205.xlsx"
        output_file = "NewsResult_with_sentiment.xlsx"
        model_path = "saved_models_20241210_004441/sentiment_model.pkl"
        vectorizer_path = "saved_models_20241210_004441/tfidf_vectorizer.pkl"

        # 기존 엑셀 파일을 원본 그대로 복사
        logger.info(f"기존 엑셀 파일 복사 중: {input_file} -> {output_file}")
        wb = openpyxl.load_workbook(input_file)
        wb.save(output_file)
        
        # 데이터 분석을 위해 pandas로 읽기
        logger.info("데이터 분석을 위해 엑셀 파일 로드 중...")
        df = pd.read_excel(input_file)
        logger.info(f"로드된 데이터 shape: {df.shape}")

        # 모델과 벡터라이저 로드
        model, vectorizer = load_model_and_vectorizer(model_path, vectorizer_path)

        # 감성 분석 수행
        logger.info("감성 분석 시작...")
        sentiment_labels = analyze_sentiments(df, model, vectorizer)

        # 새로운 워크북 열기
        wb = openpyxl.load_workbook(output_file)
        ws = wb.active   

        # 새로운 칼럼 헤더 추가
        last_col = ws.max_column
        ws.cell(row=1, column=last_col + 1, value='sentiment_label')
        ws.cell(row=1, column=last_col + 2, value='sentiment_desc')

        # 감성 분석 결과 추가
        for i, (label, desc) in enumerate(zip(sentiment_labels, 
                                            [['부정', '중립', '긍정'][int(x)+1] for x in sentiment_labels]), 
                                            start=2):
            ws.cell(row=i, column=last_col + 1, value=int(label))
            ws.cell(row=i, column=last_col + 2, value=desc)

        # 변경사항 저장
        wb.save(output_file)
        logger.info(f"분석 결과가 저장되었습니다: {output_file}")

        # # 일자별로 최대 5개의 긍정, 부정, 중립 데이터 필터링
        # logger.info("일자별로 최대 5개의 긍정, 부정, 중립 데이터 필터링 중...")
        # filtered_df = filter_top_n_by_date(df, sentiment_labels, n=5)

        # # 결과 저장
        # filtered_df.to_excel(output_file, index=False)
        # logger.info(f"필터링된 데이터가 저장되었습니다: {output_file}")

        # 결과 요약을 위해 데이터프레임으로 다시 읽기
        df_result = pd.read_excel(output_file)
        sentiment_counts = df_result['sentiment_desc'].value_counts()
        
        # 결과 요약 출력
        logger.info("\n=== 감성 분석 결과 요약 ===")
        logger.info(f"전체 뉴스 수: {len(df_result)}")
        logger.info("감성 분포:")
        for sentiment, count in sentiment_counts.items():
            percentage = (count / len(df_result)) * 100
            logger.info(f"{sentiment}: {count}건 ({percentage:.1f}%)")

    except Exception as e:
        logger.error(f"프로그램 실행 중 오류 발생: {str(e)}")
        import traceback
        logger.error(f"스택 트레이스: {traceback.format_exc()}")

if __name__ == "__main__":
    main()